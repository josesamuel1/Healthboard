import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework import viewsets, permissions
from .models import *
from .serializers import *

# Função criada para exportar dados para Excel
def export_to_excel(queryset, model_name):
    # Criando o workbook e uma sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{model_name} Data"

    # Definição dos cabeçalhos baseado nos campos do modelo
    headers = [field.verbose_name for field in queryset.model._meta.fields]
    sheet.append(headers)

    # Parte para inserir os dados
    for obj in queryset:
        data = [getattr(obj, field.name) for field in queryset.model._meta.fields]
        sheet.append(data)

    # Ajusta a largura das colunas
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 15

    # Realiza o salvamento do arquivo em uma resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={model_name}_data.xlsx'
    workbook.save(response)
    return response


# Classe atualizada e com função para exportar o arquivo
class ClassificacaoDeRiscoModelViewSet(viewsets.ModelViewSet):
    # Busca os dados no BD do model 'ClassificacaoDeRisco'
    queryset = ClassificacaoDeRisco.objects.all()

    # Serializa os dados em JSON da classe
    serializer_class = ClassificacaoDeRiscoSerializer

    # Verifica se o usuário tem as permissões concedidas no Painel do Admin
    permission_classes = [permissions.DjangoModelPermissions]
    
    def export_excel(self, request):
        queryset = self.get_queryset()
        return export_to_excel(queryset, 'ClassificacaoDeRisco')


# Classe atualizada e com função para exportar o arquivo
class BlocoCirurgicoModelViewSet(viewsets.ModelViewSet):
    # Busca os dados no BD do model 'BlocoCirurgico'
    queryset = BlocoCirurgico.objects.all()

    # Serializa os dados em JSON da classe
    serializer_class = BlocoCirurgicoSerializer

    # Verifica se o usuário tem as permissões concedidas no Painel do Admin
    permission_classes = [permissions.DjangoModelPermissions]

    def export_excel(self, request):
        queryset = self.get_queryset()
        return export_to_excel(queryset, 'BlocoCirurgico')


# Classe atualizada e com função para exportar o arquivo
class UnidadeDeInternacaoModelViewSet(viewsets.ModelViewSet):
    # Busca os dados no BD do model 'UnidadeDeInternacao'
    queryset = UnidadeDeInternacao.objects.all()

    # Serializa os dados em JSON da classe
    serializer_class = UnidadeDeInternacaoSerializer

    # Verifica se o usuário tem as permissões concedidas no Painel do Admin
    permission_classes = [permissions.DjangoModelPermissions]

    def export_excel(self, request):
        queryset = self.get_queryset()
        return export_to_excel(queryset, 'UnidadeDeInternacao')
